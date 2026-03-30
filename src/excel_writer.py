"""Запись данных в XLSX файлы."""

import os
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


class ExcelWriter:
    """Класс для записи данных в Excel."""

    @staticmethod
    def write_catalog(
            products: List[Dict[str, Any]],
            filepath: str
    ):
        """
        Записывает каталог товаров в XLSX.
        """
        print(f"\nСохраняем {len(products)} товаров в {filepath}...")

        # Создаём workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Каталог"

        if not products:
            wb.save(filepath)
            print("Файл пустой (товары не найдены)")
            return

        headers = list(products[0].keys())

        # Стиль для заголовков
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Записываем заголовки
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, size=11, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Записываем данные
        for row_num, product in enumerate(products, 2):
            for col_num, header in enumerate(headers, 1):
                value = product.get(header, '')
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        # Автоширина колонок
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)

            max_length = 0
            for row in ws[column_letter]:
                try:
                    if len(str(row.value)) > max_length:
                        max_length = len(str(row.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)  # Максимум 50
            ws.column_dimensions[column_letter].width = adjusted_width

        # Закрепляем первую строку
        ws.freeze_panes = 'A2'

        os.makedirs(os.path.dirname(filepath) or '.', exist_ok=True)
        wb.save(filepath)
        print(f"Файл успешно сохранён: {filepath}")